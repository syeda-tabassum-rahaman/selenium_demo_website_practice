import openpyxl
path = "C:\selenium_projects\data\sampledatafoodsales.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active

rows = sheet.max_row #40

cols = sheet.max_column #11

print(rows)
print(cols)

for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(row=r, column=c).value, end = "    ")

    print()


